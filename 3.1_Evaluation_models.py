import os
import pandas as pd
import numpy as np
from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing.image import load_img, img_to_array
import random
import openpyxl
from openpyxl.styles import PatternFill

# Paths to images and metadata files
images_custom_path = 'Clothing_Dataset_small/images_custom'
images_cleaned_path = 'Clothing_Dataset_small/images'
custom_metadata_path = 'custom_clothing_metadata.csv'
cleaned_metadata_path = 'cleaned_metadata.csv'

# Load custom metadata CSV file
custom_metadata = pd.read_csv(custom_metadata_path)

# Ensure 'id' and 'masterCategory' columns are treated as strings
custom_metadata['id'] = custom_metadata['id'].astype(str) + ".jpg"
custom_metadata['masterCategory'] = custom_metadata['masterCategory'].astype(str)

# Ensure all files exist and remove missing files
custom_metadata['file_exists'] = custom_metadata['id'].apply(lambda x: os.path.exists(os.path.join(images_custom_path, x)))
custom_metadata = custom_metadata[custom_metadata['file_exists']]

# Load cleaned metadata CSV file
cleaned_metadata = pd.read_csv(cleaned_metadata_path)

# Ensure 'id' and 'masterCategory' columns are treated as strings
cleaned_metadata['id'] = cleaned_metadata['id'].astype(str) + ".jpg"
cleaned_metadata['masterCategory'] = cleaned_metadata['masterCategory'].astype(str)

# Ensure all files exist and remove missing files
cleaned_metadata['file_exists'] = cleaned_metadata['id'].apply(lambda x: os.path.exists(os.path.join(images_cleaned_path, x)))
cleaned_metadata = cleaned_metadata[cleaned_metadata['file_exists']]

# Select 30 random images from cleaned metadata
random_cleaned_metadata = cleaned_metadata.sample(n=30, random_state=42)

# Load models
models = {
    'VGG16': load_model('clothing_classification_fewer_classes_model_vgg16.keras'),
    'Deep CNN': load_model('clothing_classification_model_deep_cnn.keras'),
    'Base Model': load_model('clothing_classification_model.keras'),
    'Custom VGG16': load_model('training_custom_dataset_vgg16.keras')  # New model
}

# Example class indices mapping, you need to replace this with your actual class indices
class_indices = {
    'VGG16': {'Apparel': 0, 'Footwear': 1, 'Accessories': 2, 'Personal Care': 3, 'Free Items': 4},
    'Deep CNN': {'Apparel': 0, 'Footwear': 1, 'Accessories': 2, 'Personal Care': 3, 'Free Items': 4},
    'Base Model': {'Apparel': 0, 'Footwear': 1, 'Accessories': 2, 'Personal Care': 3, 'Free Items': 4},
    'Custom VGG16': {'Apparel': 0, 'Footwear': 1, 'Accessories': 2, 'Personal Care': 3, 'Free Items': 4}  # New model
}

# Reverse class indices to get class names from predictions
rev_class_indices = {name: {v: k for k, v in indices.items()} for name, indices in class_indices.items()}

# Preprocess function
def preprocess_image(image_path, target_size=(128, 128)):
    img = load_img(image_path, target_size=target_size)
    img_array = img_to_array(img)
    img_array = np.expand_dims(img_array, axis=0)
    img_array = img_array / 255.0  # Rescale
    return img_array

# Predict function
def predict_master_category(image_path, models):
    img_array = preprocess_image(image_path)
    predictions = {}
    for model_name, model in models.items():
        preds = model.predict(img_array)
        predicted_class = np.argmax(preds, axis=1)
        predictions[model_name] = predicted_class[0]
    return predictions

# Create a DataFrame to store results
results = []

# Iterate over each image in the custom dataset
for index, row in custom_metadata.iterrows():
    image_id = row['id']
    true_master_category = row['masterCategory']
    image_path = os.path.join(images_custom_path, image_id)
    
    # Predict using each model
    predictions = predict_master_category(image_path, models)
    
    # Convert numeric predictions to class labels
    for model_name, pred in predictions.items():
        predictions[model_name] = rev_class_indices[model_name][pred]
    
    # Add results to the list
    results.append({
        'Image ID': image_id,
        'True Master Category': true_master_category,
        'VGG16 Prediction': predictions['VGG16'],
        'Deep CNN Prediction': predictions['Deep CNN'],
        'Base Model Prediction': predictions['Base Model'],
        'Custom VGG16 Prediction': predictions['Custom VGG16']  # New model
    })

# Iterate over each image in the random cleaned dataset
for index, row in random_cleaned_metadata.iterrows():
    image_id = row['id']
    true_master_category = row['masterCategory']
    image_path = os.path.join(images_cleaned_path, image_id)
    
    # Predict using each model
    predictions = predict_master_category(image_path, models)
    
    # Convert numeric predictions to class labels
    for model_name, pred in predictions.items():
        predictions[model_name] = rev_class_indices[model_name][pred]
    
    # Add results to the list
    results.append({
        'Image ID': image_id,
        'True Master Category': true_master_category,
        'VGG16 Prediction': predictions['VGG16'],
        'Deep CNN Prediction': predictions['Deep CNN'],
        'Base Model Prediction': predictions['Base Model'],
        'Custom VGG16 Prediction': predictions['Custom VGG16']  # New model
    })

# Convert results to DataFrame
results_df = pd.DataFrame(results)

# Save results to an Excel file with conditional formatting
results_excel_path = 'prediction_results.xlsx'
results_df.to_excel(results_excel_path, index=False)

# Open the saved excel file and apply conditional formatting
wb = openpyxl.load_workbook(results_excel_path)
ws = wb.active

# Define the fill for correct predictions
correct_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')

# Apply conditional formatting for each cell
for row in ws.iter_rows(min_row=2, min_col=3, max_col=6):  # Adjust according to the columns
    for cell in row:
        true_value = ws.cell(row=cell.row, column=2).value
        if cell.value == true_value:
            cell.fill = correct_fill

# Save the workbook
wb.save(results_excel_path)

# Display results
print(results_df)
